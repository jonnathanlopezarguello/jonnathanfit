from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ─── PALETA TIERRA ────────────────────────────────────────────────────────────
SAND  = "F5ECD7"; LINEN = "FAF3E8"; TERR  = "C98B6B"; CLAY  = "A0634A"
SAGE  = "9DAD8E"; WARM  = "C4A882"; RUST  = "C07B5A"; MOSS  = "7E9B7B"
DUST  = "B89F88"; TEXT_D= "3C2F22"; TEXT_L= "FFFFFF"; MUTED = "8C7B6E"
GREEN = "7A9B6A"; RED   = "C07070"

RATE  = 4400   # 1 USD ≈ 4 400 COP (junio 2026 estimado)

def cop(usd): return f"$ {usd*RATE:,.0f}" if usd else "GRATIS"
def side(s="thin", c="E8DDD0"): return Side(style=s, color=c)
def fill(h): return PatternFill("solid", fgColor=h)
def font(bold=False, sz=10, color=TEXT_D, italic=False):
    return Font(name="Calibri", bold=bold, size=sz, color=color, italic=italic)
def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def border(s="thin"):
    t = side(s); return Border(left=t, right=t, top=t, bottom=t)
def thin_border():
    return Border(left=side("thin"), right=side("thin"),
                  top=side("hair"), bottom=side("hair"))

def sec_hdr(ws, row, txt):
    ws.merge_cells(f"A{row}:F{row}")
    c = ws[f"A{row}"]; c.value = txt.upper()
    c.font = font(True, 9, TEXT_L); c.fill = fill(TERR)
    c.alignment = align("left","center"); ws.row_dimensions[row].height = 22

def hdr_row(ws, row, cols, bg=SAND):
    keys = list(cols.keys()); letters = list("ABCDEF")
    for i,(col,txt) in enumerate(cols.items()):
        c = ws[f"{col}{row}"]; c.value = txt
        c.font = font(True, 9); c.fill = fill(bg)
        c.alignment = align("center","center"); c.border = border()
    ws.row_dimensions[row].height = 18

def dat(ws, row, vals, alt=False):
    bg = SAND if alt else LINEN; cols = list("ABCDEF")[:len(vals)]
    for col, val in zip(cols, vals):
        c = ws[f"{col}{row}"]; c.value = val
        c.font = font(sz=9); c.fill = fill(bg)
        c.alignment = align("left","center",wrap=True); c.border = thin_border()
    ws.row_dimensions[row].height = 28

# ══════════════════════════════════════════════════════════════════════════════
# HOJA 1 — COSTOS EN COP + QUIÉN / CÓMO PAGAR
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active; ws1.title = "Costos COP"
ws1.sheet_view.showGridLines = False
for row in ws1.iter_rows(min_row=1, max_row=90, min_col=1, max_col=8):
    for cell in row: cell.fill = fill(LINEN)

ws1.column_dimensions["A"].width = 22
ws1.column_dimensions["B"].width = 28
ws1.column_dimensions["C"].width = 14
ws1.column_dimensions["D"].width = 16
ws1.column_dimensions["E"].width = 22
ws1.column_dimensions["F"].width = 26

# Título
ws1.merge_cells("A1:F1"); c = ws1["A1"]
c.value = "JONNATHAN FIT — Costos en Pesos Colombianos (COP)"; c.font = font(True, 15, CLAY)
c.fill = fill(SAND); c.alignment = align("left","center"); ws1.row_dimensions[1].height = 36

ws1.merge_cells("A2:F2"); c = ws1["A2"]
c.value = f"Tasa de cambio estimada: 1 USD ≈ $ {RATE:,} COP (junio 2026)  ·  Todos los precios son aproximados"
c.font = font(sz=9, color=MUTED, italic=True); c.fill = fill(SAND)
c.alignment = align("left","center"); ws1.row_dimensions[2].height = 18
ws1.row_dimensions[3].height = 8

# ── SECCIÓN 1: PAGOS A PLATAFORMAS ──
sec_hdr(ws1, 4, "1. Pagos a plataformas (costos fijos)")
hdr_row(ws1, 5, {"A":"Servicio","B":"Qué es","C":"USD","D":"COP","E":"Dónde pagar","F":"Método de pago Colombia"})
plataformas = [
    ("Supabase (plan gratuito)",
     "Backend, base de datos y autenticación",
     "GRATIS","GRATIS",
     "supabase.com",
     "No requiere pago inicial"),
    ("Google Play Console",
     "Publicar app en Play Store (pago único)",
     "USD 25", cop(25),
     "play.google.com/console",
     "Tarjeta débito/crédito Visa o Mastercard con cupo internacional (Bancolombia, Davivienda, Nequi Virtual)"),
    ("Apple Developer Program",
     "Publicar en App Store (anual)",
     "USD 99/año", cop(99)+"/año",
     "developer.apple.com",
     "Tarjeta de crédito internacional con cupo en USD (Bancolombia, Scotiabank, Falabella)"),
    ("Dominio .com o .fit",
     "Nombre propio ej. jonnathanfit.com",
     "USD 12/año", cop(12)+"/año",
     "namecheap.com / porkbun.com",
     "Tarjeta débito internacional o PayPal · También GoDaddy Colombia acepta PSE"),
    ("Supabase Pro (si escala)",
     "Más almacenamiento y funciones avanzadas",
     "USD 25/mes", cop(25)+"/mes",
     "supabase.com/pricing",
     "Tarjeta débito/crédito con cupo internacional"),
    ("Hosting adicional (AWS/GCP)",
     "Cuando la app crezca más de 500 usuarios activos",
     "USD 20–100/mes", f"$ {20*RATE:,}–{100*RATE:,}/mes",
     "aws.amazon.com / cloud.google.com",
     "Tarjeta internacional o cuenta AWS con factura en COP"),
]
for i, r in enumerate(plataformas):
    dat(ws1, 6+i, r, alt=bool(i%2))

ws1.row_dimensions[12].height = 8

# ── SECCIÓN 2: DESARROLLADOR ──
sec_hdr(ws1, 13, "2. Desarrollador React Native — opciones en Colombia y Latinoamérica")
hdr_row(ws1, 14, {"A":"Opción","B":"Perfil","C":"Tarifa/hora","D":"Costo estimado","E":"Dónde encontrar","F":"Cómo pagar"})
devs = [
    ("Freelancer Colombia Jr",
     "1–2 años experiencia React Native",
     "USD 10–15/h", f"$ {10*160*RATE:,.0f}–{15*160*RATE:,.0f}\npor mes",
     "Workana.com · Freelancer.com",
     "Transferencia Bancolombia, Nequi, Daviplata o Efecty"),
    ("Freelancer Colombia Mid",
     "3–5 años, entrega más rápido y con menos errores",
     "USD 20–30/h", f"$ {20*160*RATE:,.0f}–{30*160*RATE:,.0f}\npor mes",
     "Workana.com · LinkedIn Colombia",
     "Transferencia bancaria, Nequi o PayPal"),
    ("Agencia Colombia (MVP)",
     "Empresa local, incluye diseño y QA",
     "USD 5,000–15,000",f"$ {5000*RATE:,.0f}–\n$ {15000*RATE:,.0f}\ntotal",
     "Medellín/Bogotá Digital, Koombea, Yuxi Global",
     "Factura en COP, transferencia bancaria, pago por hitos"),
    ("Freelancer Latam Senior",
     "Argentina, México, Chile — experiencia en apps fitness",
     "USD 30–50/h", f"$ {30*160*RATE:,.0f}–{50*160*RATE:,.0f}\npor mes",
     "Toptal · Upwork · LinkedIn",
     "PayPal, Wise, o tarjeta de crédito"),
    ("Desarrollador propio",
     "Aprender React Native (cursos Udemy/Platzi)",
     "Cursos: USD 20–200",f"$ {20*RATE:,.0f}–\n$ {200*RATE:,.0f}\ncursos",
     "Platzi.com · Udemy.com",
     "Tarjeta débito/crédito · Nequi · PSE en Platzi"),
]
for i, r in enumerate(devs):
    dat(ws1, 15+i, r, alt=bool(i%2))

ws1.row_dimensions[20].height = 8

# ── SECCIÓN 3: RESUMEN TOTAL ──
sec_hdr(ws1, 21, "3. Resumen total del proyecto — presupuesto mínimo y máximo")
hdr_row(ws1, 22, {"A":"Fase","B":"Descripción","C":"Mín COP","D":"Máx COP","E":"Cuándo se paga","F":"Notas"})
resumen = [
    ("Fase 1 – Backend",    "Supabase + configuración",                  "GRATIS",                    "GRATIS",                     "Semanas 1–2",     "Solo tiempo, cero costo"),
    ("Fase 2 – Play Store", "Google Play Console + TWA",                  cop(25),                     cop(25),                      "Semanas 3–4",     "Pago único, nunca más"),
    ("Fase 3 – App nativa", "Desarrollador React Native (2 meses)",       f"$ {5000*RATE:,.0f}",       f"$ {15000*RATE:,.0f}",       "Meses 2–3",       "Pago por hitos: 30/40/30 %"),
    ("Fase 4 – Ciencia",    "Motor personalización (adicional o incluido)",f"$ {500*RATE:,.0f}",        f"$ {3000*RATE:,.0f}",        "Meses 4–5",       "Suele incluirse en Fase 3"),
    ("Fase 5 – App Store",  "Apple Developer Program",                    cop(99)+"/año",              cop(99)+"/año",               "Mes 6+",          "Renovar cada año"),
    ("Hosting recurrente",  "Supabase Pro + dominio (anual)",             f"$ {(25*12+12)*RATE:,.0f}/año",f"$ {(100*12+12)*RATE:,.0f}/año","Mensual/Anual","Crece con usuarios"),
    ("TOTAL PROYECTO",      "Lanzamiento completo (sin dev propio)",      f"$ {(25+99+500)*RATE:,.0f}",f"$ {(25+99+15000+3000)*RATE:,.0f}","—",          "Rango según desarrollador elegido"),
]
for i, r in enumerate(resumen):
    if i == len(resumen)-1:
        # fila total en negrita
        bg = TERR
        for j, (col, val) in enumerate(zip("ABCDEF", r)):
            c = ws1[f"{col}{23+i}"]; c.value = val
            c.font = font(True, 9, TEXT_L); c.fill = fill(bg)
            c.alignment = align("center","center",wrap=True); c.border = border()
        ws1.row_dimensions[23+i].height = 28
    else:
        dat(ws1, 23+i, r, alt=bool(i%2))

# ── SECCIÓN 4: MÉTODOS DE PAGO INTERNACIONALES DESDE COLOMBIA ──
ws1.row_dimensions[30].height = 8
sec_hdr(ws1, 31, "4. Cómo pagar servicios internacionales desde Colombia")
hdr_row(ws1, 32, {"A":"Método","B":"Cómo funciona","C":"Límite aprox.","D":"Costo","E":"Ideal para","F":"Banco / App"})
pagos = [
    ("Tarjeta débito virtual Nequi",
     "Genera tarjeta Visa virtual con cupo en COP para pagar en USD",
     "USD 1,500/mes","Gratis",
     "Play Store, dominios, Supabase",
     "Nequi (Bancolombia)"),
    ("Tarjeta crédito internacional",
     "Tarjeta Visa/Mastercard con cupo en USD habilitado",
     "Según cupo","Comisión ~3% TRM",
     "Todo: Apple, Google, AWS, Supabase",
     "Bancolombia, Davivienda, Scotiabank, Falabella"),
    ("Tarjeta Nu (Nubank Colombia)",
     "Tarjeta crédito Mastercard sin comisión por uso internacional",
     "Según cupo","Sin comisión",
     "Todo — la más conveniente en Colombia",
     "Nubank Colombia (app)"),
    ("Wise (Transferwise)",
     "Cuenta en USD, envía y recibe en múltiples monedas",
     "Sin límite","~0.5–1% conversión",
     "Pagar freelancers internacionales, AWS",
     "wise.com"),
    ("PayPal",
     "Cuenta vinculada a banco colombiano para pagos internacionales",
     "USD 10,000/año","~3.5% conversión",
     "Freelancers, Namecheap, Udemy",
     "paypal.com"),
    ("PSE / transferencia COP",
     "Pago directo en pesos en sitios con convenio Colombia",
     "Sin límite","Según banco",
     "GoDaddy Colombia, Platzi, algunas agencias locales",
     "Cualquier banco colombiano"),
]
for i, r in enumerate(pagos):
    dat(ws1, 33+i, r, alt=bool(i%2))


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 2 — GANTT CON PRESUPUESTO POR FASE
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Gantt")
ws2.sheet_view.showGridLines = False
for row in ws2.iter_rows(min_row=1, max_row=60, min_col=1, max_col=13):
    for cell in row: cell.fill = fill(LINEN)

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 13
ws2.column_dimensions["C"].width = 18

MONTHS = ["Sem 1–2","Sem 3–4","Mes 2","Mes 3","Mes 4","Mes 5","Mes 6","Mes 7+"]
BAR_COLS = list("DEFGHIJK")

for col in BAR_COLS:
    ws2.column_dimensions[col].width = 9

# Título
ws2.merge_cells("A1:K1"); c = ws2["A1"]
c.value = "GANTT — JONNATHAN FIT · Presupuesto por Fase (COP)"; c.font = font(True, 14, CLAY)
c.fill = fill(SAND); c.alignment = align("left","center"); ws2.row_dimensions[1].height = 34

ws2.merge_cells("A2:K2"); c = ws2["A2"]
c.value = f"Tasa estimada: 1 USD ≈ $ {RATE:,} COP  ·  Tonos tierra  ·  Mínimo viable → App Store + Play Store"
c.font = font(sz=9, color=MUTED, italic=True); c.fill = fill(SAND)
c.alignment = align("left","center"); ws2.row_dimensions[2].height = 18
ws2.row_dimensions[3].height = 8

# Encabezados
for cell_ref, txt, bg in [("A4","FASE",CLAY),("B4","PRESUPUESTO COP",CLAY),("C4","QUIÉN PAGAR",CLAY)]:
    c = ws2[cell_ref]; c.value = txt
    c.font = font(True, 9, TEXT_L); c.fill = fill(bg)
    c.alignment = align("center","center"); c.border = border()

for i,(col,month) in enumerate(zip(BAR_COLS, MONTHS)):
    c = ws2[f"{col}4"]; c.value = month
    c.font = font(True, 8); c.fill = fill(SAND)
    c.alignment = align("center","center"); c.border = border()
ws2.row_dimensions[4].height = 22

# Fases: (nombre, presupuesto_texto, quien_pagar, color, col_s, col_e)
phases = [
    ("01  Supabase + Auth\nConectar backend a la PWA actual",
     "GRATIS",
     "supabase.com\n(tarjeta, plan free)",
     SAGE, "D", "F"),
    ("02  Play Store vía TWA\nPublicar sin reescribir código",
     cop(25)+"\n(pago único)",
     "Google Play Console\nTarjeta Nequi / Nubank",
     WARM, "F", "H"),
    ("03  React Native\nApp nativa iOS + Android",
     f"$ {5000*RATE:,.0f} – $ {15000*RATE:,.0f}\n(dev 2 meses · pago por hitos)",
     "Workana.com / Agencia\nTransferencia bancaria",
     RUST, "H", "J"),
    ("04  Motor Científico\nPersonalización por usuario",
     f"$ {500*RATE:,.0f} – $ {3000*RATE:,.0f}\n(incluido en dev o adicional)",
     "Mismo desarrollador\n+ Claude API (Anthropic)",
     MOSS, "J", "K"),
    ("05  App Store iOS\nRevisión + publicación Apple",
     cop(99)+"/año\n(cuenta Apple Developer)",
     "developer.apple.com\nTarjeta crédito internacional",
     DUST, "K", "K"),
]

for ri, (name, budget, who, bar_color, col_s, col_e) in enumerate(phases):
    row = 5 + ri * 4
    alt_bg = SAND if ri % 2 == 0 else LINEN

    start_i = BAR_COLS.index(col_s) if col_s in BAR_COLS else 0
    end_i   = (BAR_COLS.index(col_e) + 1) if col_e in BAR_COLS else len(BAR_COLS)

    # Nombre fase (3 rows de alto)
    ws2.merge_cells(f"A{row}:A{row+2}")
    c = ws2[f"A{row}"]; c.value = name
    c.font = font(True, 9); c.fill = fill(alt_bg)
    c.alignment = align("left","center",wrap=True)
    c.border = Border(left=side("thin"), right=side("thin"),
                      top=side("thin"), bottom=side("thin"))

    # Presupuesto
    ws2.merge_cells(f"B{row}:B{row+2}")
    c = ws2[f"B{row}"]; c.value = budget
    c.font = font(True, 9, CLAY); c.fill = fill(alt_bg)
    c.alignment = align("center","center",wrap=True)
    c.border = Border(left=side("thin"), right=side("thin"),
                      top=side("thin"), bottom=side("thin"))

    # Quién pagar
    ws2.merge_cells(f"C{row}:C{row+2}")
    c = ws2[f"C{row}"]; c.value = who
    c.font = font(sz=8, color=MUTED, italic=True); c.fill = fill(alt_bg)
    c.alignment = align("left","center",wrap=True)
    c.border = Border(left=side("thin"), right=side("thin"),
                      top=side("thin"), bottom=side("thin"))

    # Barras Gantt (3 filas de altura)
    for ci, col in enumerate(BAR_COLS):
        in_bar = start_i <= ci < end_i
        for sub in range(3):
            c = ws2[f"{col}{row+sub}"]
            c.fill = fill(bar_color if in_bar else alt_bg)
            c.border = Border(
                top=side("hair", bar_color if in_bar else "E8DDD0"),
                bottom=side("hair", bar_color if in_bar else "E8DDD0"),
                left=side("thin" if (in_bar and ci==start_i) else "hair",
                          bar_color if in_bar else "E8DDD0"),
                right=side("thin" if (in_bar and ci==end_i-1) else "hair",
                           bar_color if in_bar else "E8DDD0"))

    # Etiqueta en barra (fila del medio)
    label_col_s = col_s if col_s in BAR_COLS else BAR_COLS[0]
    label_col_e = BAR_COLS[min(end_i-1, len(BAR_COLS)-1)]
    if label_col_s != label_col_e:
        ws2.merge_cells(f"{label_col_s}{row+1}:{label_col_e}{row+1}")
    c = ws2[f"{label_col_s}{row+1}"]
    c.value = name.split("\n")[0].split("  ")[1] if "  " in name else name.split("\n")[0]
    c.font = font(True, 8, TEXT_L); c.fill = fill(bar_color)
    c.alignment = align("center","center")

    ws2.row_dimensions[row].height   = 18
    ws2.row_dimensions[row+1].height = 18
    ws2.row_dimensions[row+2].height = 18
    ws2.row_dimensions[row+3].height = 6

# ── FILA TOTALES ──
total_row = 5 + len(phases)*4 + 1
ws2.merge_cells(f"A{total_row}:K{total_row}")
c = ws2[f"A{total_row}"]
c.value = (f"TOTAL ESTIMADO:  mínimo $ {(25+99+500)*RATE:,.0f} COP  ·  "
           f"máximo $ {(25+99+15000+3000)*RATE:,.0f} COP  "
           f"(sin contar hosting mensual recurrente)")
c.font = font(True, 10, TEXT_L); c.fill = fill(CLAY)
c.alignment = align("center","center"); c.border = border()
ws2.row_dimensions[total_row].height = 26

# ── LEYENDA COLORES ──
leg_row = total_row + 2
ws2.merge_cells(f"A{leg_row}:C{leg_row}")
c = ws2[f"A{leg_row}"]; c.value = "LEYENDA"
c.font = font(True, 8, TEXT_L); c.fill = fill(TERR)
c.alignment = align("left","center"); ws2.row_dimensions[leg_row].height = 18

legend = [(SAGE,"Fase 1 – Gratis"),(WARM,"Fase 2 – $ 110.000"),(RUST,"Fase 3 – $ 22M – 66M"),(MOSS,"Fase 4 – $ 2,2M – 13,2M"),(DUST,"Fase 5 – $ 435.600/año")]
for idx,(color,lbl) in enumerate(legend):
    c1 = ws2[f"{BAR_COLS[idx*0]}{leg_row+1}"] if idx==0 else ws2[f"D{leg_row+1}"]
    # simplificado: usar columnas D..K
    bcol = BAR_COLS[idx] if idx < len(BAR_COLS) else "K"
    c = ws2[f"{bcol}{leg_row+1}"]
    c.value = lbl; c.font = font(True, 8, TEXT_L)
    c.fill = fill(color); c.alignment = align("center","center"); c.border = border()
ws2.row_dimensions[leg_row+1].height = 18

# ── NOTA ──
note_row = leg_row + 3
ws2.merge_cells(f"A{note_row}:K{note_row}")
c = ws2[f"A{note_row}"]
c.value = ("Nota: Tasa de cambio referencial (no financiero). Fase 1 puede comenzar de inmediato. "
           "Para ahorrar, se recomienda contratar desarrollador en Colombia vía Workana o estudiar React Native en Platzi/Udemy. "
           "Nubank Colombia es la mejor opción para pagos internacionales sin comisión.")
c.font = font(sz=8, color=MUTED, italic=True); c.fill = fill(LINEN)
c.alignment = align("left","center",wrap=True); ws2.row_dimensions[note_row].height = 32

path = "/home/user/jonnathanfit/JonnathanFit_HojaDeRuta_COP.xlsx"
wb.save(path)
print(f"Guardado: {path}")
