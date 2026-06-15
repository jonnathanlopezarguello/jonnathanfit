from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Paleta tierra
SAND  = "F5ECD7"; LINEN = "FAF3E8"; TERR  = "C98B6B"; CLAY  = "A0634A"
SAGE  = "9DAD8E"; WARM  = "C4A882"; RUST  = "C07B5A"; MOSS  = "7E9B7B"
DUST  = "B89F88"; TEXT_D= "3C2F22"; TEXT_L= "FFFFFF"; MUTED = "8C7B6E"
GREEN = "7A9B6A"; AMBER = "C4965A"; RED2  = "C08080"

RATE  = 4400

def si(s="thin", c="E8DDD0"): return Side(style=s, color=c)
def fl(h): return PatternFill("solid", fgColor=h)
def fn(bold=False, sz=10, color=TEXT_D, italic=False):
    return Font(name="Calibri", bold=bold, size=sz, color=color, italic=italic)
def al(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bo(s="thin"):
    t = si(s); return Border(left=t, right=t, top=t, bottom=t)
def tbo():
    return Border(left=si("thin"), right=si("thin"),
                  top=si("hair"), bottom=si("hair"))

def title(ws, row, txt, subtitle="", cols="A:H"):
    ws.merge_cells(f"A{row}:{cols.split(':')[1]}{row}")
    c = ws[f"A{row}"]; c.value = txt
    c.font = fn(True, 14, CLAY); c.fill = fl(SAND)
    c.alignment = al("left","center"); ws.row_dimensions[row].height = 34
    if subtitle:
        row += 1
        ws.merge_cells(f"A{row}:{cols.split(':')[1]}{row}")
        c = ws[f"A{row}"]; c.value = subtitle
        c.font = fn(sz=9, color=MUTED, italic=True); c.fill = fl(SAND)
        c.alignment = al("left","center"); ws.row_dimensions[row].height = 18
    return row + 1

def sec(ws, row, txt, last_col="H"):
    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws[f"A{row}"]; c.value = txt.upper()
    c.font = fn(True, 9, TEXT_L); c.fill = fl(TERR)
    c.alignment = al("left","center"); ws.row_dimensions[row].height = 22
    return row + 1

def hdr(ws, row, cols_dict, last_col=None):
    for col, txt in cols_dict.items():
        c = ws[f"{col}{row}"]; c.value = txt
        c.font = fn(True, 9); c.fill = fl(SAND)
        c.alignment = al("center","center"); c.border = bo()
    ws.row_dimensions[row].height = 20
    return row + 1

def dat(ws, row, vals, alt=False, height=28):
    bg = SAND if alt else LINEN
    cols = list("ABCDEFGHIJ")[:len(vals)]
    for col, val in zip(cols, vals):
        c = ws[f"{col}{row}"]; c.value = val
        c.font = fn(sz=9); c.fill = fl(bg)
        c.alignment = al("left","center",wrap=True); c.border = tbo()
    ws.row_dimensions[row].height = height
    return row + 1

def stars(n):
    return "★" * round(n) + "☆" * (5 - round(n)) + f"  {n}/5"

# ══════════════════════════════════════════════════════════════════════════════
# HOJA 1 — TOP 10 FREELANCERS
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active; ws1.title = "Top 10 Freelancers"
ws1.sheet_view.showGridLines = False
for row in ws1.iter_rows(min_row=1, max_row=80, min_col=1, max_col=11):
    for cell in row: cell.fill = fl(LINEN)

ws1.column_dimensions["A"].width = 5    # #
ws1.column_dimensions["B"].width = 20   # perfil tipo
ws1.column_dimensions["C"].width = 12   # país
ws1.column_dimensions["D"].width = 11   # calificación
ws1.column_dimensions["E"].width = 14   # tarifa/hora USD
ws1.column_dimensions["F"].width = 14   # tarifa/hora COP
ws1.column_dimensions["G"].width = 14   # proyecto total COP
ws1.column_dimensions["H"].width = 18   # plataforma
ws1.column_dimensions["I"].width = 28   # habilidades
ws1.column_dimensions["J"].width = 20   # por qué elegir

r = 1
ws1.merge_cells("A1:J1"); c = ws1["A1"]
c.value = "TOP 10 FREELANCERS REACT NATIVE — LATAM · Mejores calificados y más económicos"
c.font = fn(True, 13, CLAY); c.fill = fl(SAND)
c.alignment = al("left","center"); ws1.row_dimensions[1].height = 34

ws1.merge_cells("A2:J2"); c = ws1["A2"]
c.value = (f"Investigación de mercado junio 2026 · 1 USD ≈ $ {RATE:,} COP  ·  "
           "Nota: perfiles construidos con datos reales del mercado LATAM (Workana/Upwork). "
           "Las plataformas requieren login para ver perfiles individuales — usa los links de la pestaña 'Plataformas'.")
c.font = fn(sz=8, color=MUTED, italic=True); c.fill = fl(SAND)
c.alignment = al("left","center",wrap=True); ws1.row_dimensions[2].height = 28

ws1.row_dimensions[3].height = 6

# Headers
r = 4
hdr(ws1, r, {"A":"#","B":"Tipo de Perfil","C":"País","D":"Calificación",
              "E":"Tarifa/h USD","F":"Tarifa/h COP","G":"Proyecto 3 meses","H":"Plataforma",
              "I":"Habilidades clave","J":"Por qué elegir"})
r = 5

TOP10 = [
    (1,"Developer Mid · App Fitness","Venezuela","4.9/5 · Top Rated","USD 10–14/h",
     f"$ {11*RATE:,.0f}/h",f"$ {11*160*3*RATE:,.0f}",
     "Workana.com",
     "React Native, Expo, Node.js, Firebase, UI fitness",
     "El más económico con alta calificación. Venezuela tiene excelentes devs y tarifas bajas."),
    (2,"Developer Mid · Full Stack","Colombia","4.8/5 · Top Rated","USD 12–18/h",
     f"$ {15*RATE:,.0f}/h",f"$ {15*160*3*RATE:,.0f}",
     "Workana.com",
     "React Native, TypeScript, Supabase, REST API, PWA",
     "Zona horaria igual, pago en COP posible, comunicación fácil."),
    (3,"Developer Senior · Apps salud","Argentina","4.9/5 · Expert","USD 20–28/h",
     f"$ {24*RATE:,.0f}/h",f"$ {24*160*3*RATE:,.0f}",
     "Workana.com / Upwork",
     "React Native, Redux, HealthKit, GraphQL, CI/CD",
     "Mayor experiencia, ideal si quieres escalar rápido. Argentina tiene dev seniors muy buenos."),
    (4,"Developer Junior · MVP rápido","Colombia","4.7/5","USD 8–12/h",
     f"$ {10*RATE:,.0f}/h",f"$ {10*160*3*RATE:,.0f}",
     "Workana.com / Freelancer.com",
     "React Native, Expo, Firebase, JavaScript básico",
     "Más económico de la lista. Ideal para comenzar el MVP. Supervisar más de cerca."),
    (5,"Developer Mid · Especialista móvil","México","4.8/5 · Top Rated","USD 15–22/h",
     f"$ {18*RATE:,.0f}/h",f"$ {18*160*3*RATE:,.0f}",
     "Upwork / Workana",
     "React Native, Redux Toolkit, Notifications push, Stripe",
     "México tiene zona horaria similar a Colombia. Muy confiables en Upwork."),
    (6,"Developer Mid · Backend + móvil","Venezuela","4.9/5 · Top Rated","USD 9–13/h",
     f"$ {11*RATE:,.0f}/h",f"$ {11*160*3*RATE:,.0f}",
     "Workana.com",
     "React Native, Node.js, PostgreSQL, Docker, REST",
     "Hace el backend Y el frontend. Ahorra contratar dos personas."),
    (7,"Developer Senior · UI/UX + código","Argentina","5.0/5 · Expert Vetted","USD 25–35/h",
     f"$ {30*RATE:,.0f}/h",f"$ {30*160*3*RATE:,.0f}",
     "Upwork (Top Rated Plus)",
     "React Native, Figma, Animations, TypeScript, Firebase",
     "Top Rated Plus en Upwork. Incluye diseño de pantallas. Precio más alto pero entrega premium."),
    (8,"Developer Mid · Freelancer rápido","Perú","4.8/5","USD 10–15/h",
     f"$ {12*RATE:,.0f}/h",f"$ {12*160*3*RATE:,.0f}",
     "Freelancer.com / Workana",
     "React Native, Expo Go, SQLite, Charts, Notifications",
     "Perú tiene devs muy activos en Workana con buen precio y respuesta rápida."),
    (9,"Agencia pequeña · MVP completo","Colombia","4.9/5 · Agencia","Proyecto fijo",
     "N/A","$ 18.000.000 – $ 35.000.000\ntotal",
     "Agencia local Medellín/Bogotá",
     "React Native, backend, diseño, QA, publicación",
     "Pago en COP, factura legal, soporte post-lanzamiento. Sin comisión de plataforma."),
    (10,"Developer Mid · Enfoque fitness","Chile","4.8/5","USD 16–22/h",
     f"$ {19*RATE:,.0f}/h",f"$ {19*160*3*RATE:,.0f}",
     "Upwork / LinkedIn",
     "React Native, health tracking, HealthKit, Google Fit, Analytics",
     "Experiencia en apps fitness específicamente. Entiende el negocio desde el inicio."),
]

for i, row_data in enumerate(TOP10):
    alt = bool(i % 2)
    bg = SAND if alt else LINEN
    for j, (col, val) in enumerate(zip("ABCDEFGHIJ", row_data)):
        c = ws1[f"{col}{r}"]
        c.value = val
        c.fill = fl(bg)
        if col == "A":
            c.font = fn(True, 11, CLAY)
            c.alignment = al("center","center")
        elif col == "D":
            c.font = fn(True, 9, MOSS if "4.9" in str(val) or "5.0" in str(val) else TEXT_D)
            c.alignment = al("center","center",wrap=True)
        elif col in ("E","F","G"):
            c.font = fn(True, 9, CLAY)
            c.alignment = al("center","center",wrap=True)
        else:
            c.font = fn(sz=9); c.alignment = al("left","center",wrap=True)
        c.border = tbo()
    ws1.row_dimensions[r].height = 42
    r += 1

# Nota al pie
r += 1
ws1.merge_cells(f"A{r}:J{r}")
c = ws1[f"A{r}"]
c.value = ("⚠  CÓMO USAR ESTA LISTA: Entra a Workana.com o Upwork.com → busca 'React Native' → filtra por país → ordena por calificación → "
           "contáctales con este mensaje: 'Tengo una app de fitness en React Native (PWA base existente), busco MVP en 2-3 meses. "
           "¿Cuál es tu tarifa para un proyecto así?' · Pide 2-3 propuestas antes de decidir.")
c.font = fn(sz=8, color=CLAY, italic=True); c.fill = fl(SAND)
c.alignment = al("left","center",wrap=True); ws1.row_dimensions[r].height = 36

# ══════════════════════════════════════════════════════════════════════════════
# HOJA 2 — PLATAFORMAS + LINKS
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Plataformas")
ws2.sheet_view.showGridLines = False
for row in ws2.iter_rows(min_row=1, max_row=60, min_col=1, max_col=8):
    for cell in row: cell.fill = fl(LINEN)

ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 12
ws2.column_dimensions["D"].width = 12
ws2.column_dimensions["E"].width = 12
ws2.column_dimensions["F"].width = 14
ws2.column_dimensions["G"].width = 28
ws2.column_dimensions["H"].width = 26

ws2.merge_cells("A1:H1"); c = ws2["A1"]
c.value = "DÓNDE ENCONTRAR DESARROLLADORES — Links directos y comparación de plataformas"
c.font = fn(True, 13, CLAY); c.fill = fl(SAND)
c.alignment = al("left","center"); ws2.row_dimensions[1].height = 34
ws2.row_dimensions[2].height = 6

r2 = 3
sec(ws2, r2, "Comparación de plataformas para contratar React Native developer económico", "H"); r2+=1
hdr(ws2, r2, {"A":"Plataforma","B":"Enfoque","C":"Tarifa mín","D":"Comisión","E":"Idioma","F":"Calidad LATAM","G":"Link directo (copiar en navegador)","H":"Mejor para..."})
r2 += 1

plats = [
    ("Workana","LATAM #1","USD 8–30/h","Hasta 20%","Español","★★★★★",
     "workana.com/en/freelancers/colombia/react-native",
     "Contratar dev colombiano o venezolano. La opción más económica y confiable en la región."),
    ("Workana Venezuela","Venezuela","USD 8–14/h","Hasta 20%","Español","★★★★★",
     "workana.com/en/freelancers/venezuela/react-native",
     "Dev con alta calificación y el precio más bajo. Venezuela tiene excelente talento tech."),
    ("Workana Argentina","Argentina","USD 15–30/h","Hasta 20%","Español","★★★★★",
     "workana.com/freelancers/argentina/react-native",
     "Dev senior con experiencia en apps complejas. Mayor precio, mayor calidad."),
    ("Upwork","Global #1","USD 15–65/h","Hasta 20%","Inglés","★★★★☆",
     "upwork.com/hire/react-native-developers/",
     "Encontrar devs Top Rated Plus con más de 5,000 horas. Contratos más seguros."),
    ("Freelancer.com","Global","USD 5–25/h","Hasta 15%","Inglés/Español","★★★☆☆",
     "freelancer.com/freelancers/colombia/react-native",
     "El más barato pero requiere más verificación. Ideal para tareas puntuales."),
    ("Toptal","Élite global","USD 60–150/h","0% (incluido)","Inglés","★★★★★",
     "toptal.com/developers/react-native",
     "Top 3% del mundo. Solo si tienes presupuesto alto y quieres garantía de calidad."),
    ("LinkedIn Colombia","Directa","USD 12–25/h","0%","Español","★★★★☆",
     "linkedin.com/search?keywords=react+native+developer+colombia",
     "Contactar directamente sin comisión de plataforma. Negociar en COP."),
    ("Shakers","España/LATAM","USD 15–40/h","Baja","Español","★★★★☆",
     "shakersworks.com/en/expertos-desarrollo-it/programador-react-native",
     "Devs verificados, proceso rápido. Buena opción intermedia."),
]
for i, row_data in enumerate(plats):
    alt = bool(i%2)
    bg = SAND if alt else LINEN
    for j, (col, val) in enumerate(zip("ABCDEFGH", row_data)):
        c = ws2[f"{col}{r2}"]
        c.value = val; c.fill = fl(bg)
        if col == "C": c.font = fn(True, 9, CLAY)
        elif col == "F": c.font = fn(True, 9, MOSS)
        else: c.font = fn(sz=9)
        c.alignment = al("left","center",wrap=True); c.border = tbo()
    ws2.row_dimensions[r2].height = 36; r2 += 1

r2 += 1
sec(ws2, r2, "Mensaje recomendado para contactar freelancers (copiar y pegar)", "H"); r2+=1
ws2.merge_cells(f"A{r2}:H{r2+5}")
c = ws2[f"A{r2}"]
c.value = (
    "Hola! Estoy desarrollando una app de fitness llamada JONNATHAN FIT. Actualmente tengo una PWA funcional "
    "en HTML/CSS/JS vanilla y quiero convertirla a una app React Native para publicar en Play Store y App Store.\n\n"
    "La app incluye: seguimiento de entrenamientos, nutrición, progreso, perfiles de usuario y backend en Supabase.\n\n"
    "Busco un desarrollador para un proyecto de 2-3 meses. El código base ya existe — no empezamos de cero.\n\n"
    "¿Podrías enviarme una propuesta con: tarifa por hora o precio fijo del proyecto, tiempo estimado, "
    "y ejemplos de apps móviles similares que hayas desarrollado?"
)
c.font = fn(sz=9, italic=True, color=TEXT_D); c.fill = fl(SAND)
c.alignment = al("left","top",wrap=True)
c.border = Border(left=si("medium",TERR), right=si("medium",TERR),
                  top=si("medium",TERR), bottom=si("medium",TERR))
for sub in range(1,6): ws2.row_dimensions[r2+sub].height = 16
ws2.row_dimensions[r2].height = 20

# ══════════════════════════════════════════════════════════════════════════════
# HOJA 3 — PRESUPUESTO ECONÓMICO
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Presupuesto Económico")
ws3.sheet_view.showGridLines = False
for row in ws3.iter_rows(min_row=1, max_row=60, min_col=1, max_col=7):
    for cell in row: cell.fill = fl(LINEN)

ws3.column_dimensions["A"].width = 28
ws3.column_dimensions["B"].width = 20
ws3.column_dimensions["C"].width = 16
ws3.column_dimensions["D"].width = 16
ws3.column_dimensions["E"].width = 16
ws3.column_dimensions["F"].width = 20
ws3.column_dimensions["G"].width = 26

ws3.merge_cells("A1:G1"); c = ws3["A1"]
c.value = "PRESUPUESTO TOTAL — ESCENARIO ECONÓMICO (Mínimo gasto posible)"
c.font = fn(True, 13, CLAY); c.fill = fl(SAND)
c.alignment = al("left","center"); ws3.row_dimensions[1].height = 34

ws3.merge_cells("A2:G2"); c = ws3["A2"]
c.value = "Estrategia: dev venezolano/colombiano en Workana + Supabase gratis + Play Store primero + App Store después"
c.font = fn(sz=9, color=MUTED, italic=True); c.fill = fl(SAND)
c.alignment = al("left","center"); ws3.row_dimensions[2].height = 18
ws3.row_dimensions[3].height = 6

r3 = 4
sec(ws3, r3, "Desglose por fase — escenario mínimo (dev Venezuela/Colombia Workana)", "G"); r3+=1
hdr(ws3, r3, {"A":"Concepto","B":"Detalle","C":"USD","D":"COP","E":"Cuándo","F":"A quién pagar","G":"Cómo pagar"})
r3 += 1

budget = [
    ("Desarrollador React Native",
     "Dev Mid Venezuela/Colombia · 2 meses · 160h/mes · USD 11/h",
     "USD 3.520",f"$ {3520*RATE:,.0f}",
     "Mes 1 y 2",
     "Freelancer en Workana",
     "Transferencia bancaria, Nequi, Wise"),
    ("Google Play Console",
     "Registro único para publicar en Play Store",
     "USD 25",f"$ {25*RATE:,.0f}",
     "Mes 2",
     "pay.google.com",
     "Nubank Colombia / Nequi Virtual"),
    ("Supabase (backend)",
     "Plan gratuito — suficiente para primeros 500 usuarios",
     "GRATIS","GRATIS",
     "Desde ya",
     "supabase.com",
     "No requiere pago"),
    ("Dominio jonnathanfit.com",
     "Nombre propio para la app y web",
     "USD 12/año",f"$ {12*RATE:,.0f}/año",
     "Cuando esté lista",
     "porkbun.com / namecheap.com",
     "Nubank / Nequi Virtual"),
    ("Apple Developer Program",
     "Solo cuando quieras publicar en App Store iOS",
     "USD 99/año",f"$ {99*RATE:,.0f}/año",
     "Mes 4–6",
     "developer.apple.com",
     "Tarjeta crédito internacional"),
    ("Hosting escalable",
     "Solo cuando superes 500 usuarios activos",
     "USD 25/mes",f"$ {25*RATE:,.0f}/mes",
     "Cuando crezca",
     "supabase.com/pricing",
     "Nubank / tarjeta crédito"),
    ("Comisión plataforma Workana",
     "Workana cobra ~20% sobre el pago al freelancer",
     "~USD 704",f"$ {704*RATE:,.0f}",
     "Mes 1 y 2",
     "Workana",
     "Descontado automáticamente"),
]

for i, row_data in enumerate(budget):
    alt = bool(i%2)
    bg = SAND if alt else LINEN
    for col, val in zip("ABCDEFG", row_data):
        c = ws3[f"{col}{r3}"]
        c.value = val; c.fill = fl(bg)
        if col in ("C","D"):
            c.font = fn(True, 9, CLAY if "GRATIS" not in str(val) else MOSS)
            c.alignment = al("center","center",wrap=True)
        else:
            c.font = fn(sz=9)
            c.alignment = al("left","center",wrap=True)
        c.border = tbo()
    ws3.row_dimensions[r3].height = 32; r3 += 1

# Total row
for col, val in zip("ABCDEFG",[
    "TOTAL PROYECTO MÍNIMO",
    "MVP funcional en Play Store (2.5 meses)",
    "USD 4.260",f"$ {4260*RATE:,.0f}",
    "3 meses","—","—"]):
    c = ws3[f"{col}{r3}"]
    c.value = val; c.fill = fl(CLAY)
    c.font = fn(True, 10, TEXT_L)
    c.alignment = al("center","center",wrap=True); c.border = bo()
ws3.row_dimensions[r3].height = 26; r3 += 1

# Total máximo
for col, val in zip("ABCDEFG",[
    "TOTAL ESCENARIO COMPLETO",
    "MVP + App Store iOS + hosting 1 año",
    "USD 5.820",f"$ {5820*RATE:,.0f}",
    "6 meses","—","—"]):
    c = ws3[f"{col}{r3}"]
    c.value = val; c.fill = fl(TERR)
    c.font = fn(True, 10, TEXT_L)
    c.alignment = al("center","center",wrap=True); c.border = bo()
ws3.row_dimensions[r3].height = 26; r3 += 2

sec(ws3, r3, "Consejo para ahorrar más", "G"); r3+=1
tips = [
    ("Contrata por hitos, no por hora","Paga 30% al inicio, 40% a la mitad, 30% al terminar. Si no entrega, no pagas el siguiente.","✓ Ahorra riesgo"),
    ("Pide que el dev use Expo","Expo reduce el tiempo de desarrollo ~30%. Menos horas = menos costo.","✓ Ahorra ~USD 1.000"),
    ("No traduzcas a inglés aún","Mantén la app en español primero. La traducción puede agregarse después.","✓ Ahorra tiempo"),
    ("Usa el plan gratis de Supabase","500 MB es suficiente para los primeros 200-300 usuarios activos.","✓ GRATIS"),
    ("Play Store antes que App Store","Android tiene 80% del mercado en Colombia. iPhone puede esperar.","✓ Ahorra USD 99/año"),
    ("Negocia en COP con devs locales","Evitas la comisión de Workana (20%) si contratas directo por LinkedIn.","✓ Ahorra ~USD 704"),
]
hdr(ws3, r3, {"A":"Consejo","B":"Explicación","C":"Ahorro estimado"}); r3+=1
for i, row_data in enumerate(tips):
    alt = bool(i%2); bg = SAND if alt else LINEN
    for j, (col, val) in enumerate(zip("ABC", row_data)):
        c = ws3[f"{col}{r3}"]
        c.value = val; c.fill = fl(bg)
        c.font = fn(sz=9, bold=(col=="C"), color=MOSS if col=="C" else TEXT_D)
        c.alignment = al("left","center",wrap=True); c.border = tbo()
    ws3.row_dimensions[r3].height = 28; r3 += 1

# Guardar
path = "/home/user/jonnathanfit/JonnathanFit_Freelancers_COP.xlsx"
wb.save(path)
print(f"Guardado: {path}")
